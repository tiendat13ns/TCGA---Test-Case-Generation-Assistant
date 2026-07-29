import io
from uuid import UUID
from fastapi import APIRouter, HTTPException, Response, Query
from sqlalchemy.exc import SQLAlchemyError
from datetime import datetime, timezone
import logging

from app.database import SessionLocal
from app.models import TestCase, Requirement
from app.schemas.test_case_schema import (
    StudioTestCaseItem,
    StudioTestCaseListResponse,
    TestCaseUpdatePayload,
    TestCaseCreatePayload,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/test-cases", tags=["test-case-studio"])

def _to_studio_item(tc: TestCase, req: Requirement) -> StudioTestCaseItem:
    return StudioTestCaseItem(
        id=str(tc.id),
        requirement_id=str(tc.requirement_id),
        document_id=str(tc.document_id) if tc.document_id else None,
        title=tc.title,
        scenario=tc.scenario,
        preconditions=tc.preconditions,
        test_steps=tc.test_steps,
        test_data=tc.test_data,
        expected_result=tc.expected_result,
        priority=tc.priority,
        severity=tc.severity,
        test_type=tc.test_type,
        automation_candidate=tc.automation_candidate,
        execution_type=tc.execution_type,
        execution_status=tc.execution_status,
        status=tc.status,
        version=tc.version,
        feature_name=req.feature_name,
        requirement_title=req.title,
        project_id=str(req.project_id) if req.project_id else None,
        module_name=req.module_name,
    )

@router.get("/export",
    responses={
        200: {
            "content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}},
            "description": "Exported Excel file containing test cases.",
        }
    },
)
def export_test_cases_studio(project_id: str | None = None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        with SessionLocal() as db:
            query = db.query(TestCase, Requirement).join(Requirement, TestCase.requirement_id == Requirement.id)
            if project_id:
                try:
                    project_uuid = UUID(project_id)
                    query = query.filter(Requirement.project_id == project_uuid)
                except ValueError:
                    raise HTTPException(status_code=400, detail="Invalid project_id format")
            
            results = query.all()
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        headers = ["Feature", "Test Case ID", "Title", "Precondition", "Test Steps", "Test Data", "Expected Output", "Priority", "Status", "Test Type"]
        ws.append(headers)
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        for idx, (tc, req) in enumerate(results):
            feature_name = req.feature_name or req.module_name or req.title or "Unknown Feature"
            steps_text = "\n".join([f"{i+1}. {s}" for i, s in enumerate(tc.test_steps)]) if tc.test_steps else ""
            
            ws.append([
                feature_name,
                f"TC-{str(idx+1).zfill(2)}",
                tc.title,
                tc.preconditions or "",
                steps_text,
                tc.test_data or "",
                tc.expected_result,
                tc.priority,
                tc.status,
                tc.test_type or ""
            ])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"test_cases_{project_id[:8]}.xlsx" if project_id else "test_cases_all.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("", response_model=StudioTestCaseListResponse)
def get_studio_test_cases(
    project_id: str | None = Query(None),
    document_id: str | None = Query(None),
    priority: str | None = Query(None),
    status: str | None = Query(None),
    test_type: str | None = Query(None)
):
    try:
        with SessionLocal() as db:
            query = db.query(TestCase, Requirement).join(Requirement, TestCase.requirement_id == Requirement.id)
            
            if project_id:
                try:
                    query = query.filter(Requirement.project_id == UUID(project_id))
                except ValueError:
                    raise HTTPException(status_code=400, detail="Invalid project_id format")
            if document_id:
                try:
                    query = query.filter(TestCase.document_id == UUID(document_id))
                except ValueError:
                    raise HTTPException(status_code=400, detail="Invalid document_id format")
            if priority:
                query = query.filter(TestCase.priority == priority)
            if status:
                query = query.filter(TestCase.status == status)
            if test_type:
                query = query.filter(TestCase.test_type == test_type)
            
            results = query.all()
            items = [_to_studio_item(tc, req) for tc, req in results]
            
            return StudioTestCaseListResponse(
                total_test_cases=len(items),
                test_cases=items
            )
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        logger.error(f"DB Error: {exc}")
        raise HTTPException(status_code=500, detail="Database error while fetching test cases") from exc


@router.put("/{test_case_id}", response_model=StudioTestCaseItem)
def update_studio_test_case(test_case_id: str, payload: TestCaseUpdatePayload):
    try:
        tc_uuid = UUID(test_case_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid test_case_id format")
        
    try:
        with SessionLocal() as db:
            tc = db.query(TestCase).filter(TestCase.id == tc_uuid).first()
            if not tc:
                raise HTTPException(status_code=404, detail="Test case not found")
                
            update_data = payload.model_dump(exclude_unset=True)
            if update_data:
                for key, value in update_data.items():
                    setattr(tc, key, value)
                tc.updated_at = datetime.now(timezone.utc)
                tc.version += 1
                
                db.commit()
                db.refresh(tc)
                
            req = db.query(Requirement).filter(Requirement.id == tc.requirement_id).first()
            if not req:
                raise HTTPException(status_code=500, detail="Requirement not found for test case")
                
            return _to_studio_item(tc, req)
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        logger.error(f"DB Error: {exc}")
        raise HTTPException(status_code=500, detail="Database error while updating test case") from exc

@router.post("", response_model=StudioTestCaseItem)
def create_studio_test_case(payload: TestCaseCreatePayload):
    try:
        req_uuid = UUID(payload.requirement_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid requirement_id format")
        
    try:
        with SessionLocal() as db:
            req = db.query(Requirement).filter(Requirement.id == req_uuid).first()
            if not req:
                raise HTTPException(status_code=404, detail="Requirement not found")
                
            tc = TestCase(
                requirement_id=req_uuid,
                document_id=req.document_id,
                title=payload.title,
                scenario=payload.scenario,
                preconditions=payload.preconditions,
                test_steps=payload.test_steps,
                test_data=payload.test_data,
                expected_result=payload.expected_result,
                priority=payload.priority,
                severity=payload.severity,
                test_type=payload.test_type,
                automation_candidate=payload.automation_candidate,
                execution_type=payload.execution_type,
                execution_status=payload.execution_status,
                status=payload.status,
            )
            db.add(tc)
            db.commit()
            db.refresh(tc)
            
            return _to_studio_item(tc, req)
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        logger.error(f"DB Error: {exc}")
        raise HTTPException(status_code=500, detail="Database error while creating test case") from exc

from pydantic import BaseModel
class BugReportPayload(BaseModel):
    actual_result: str

@router.post("/{test_case_id}/bug-report")
def generate_auto_bug_report(test_case_id: str, payload: BugReportPayload):
    try:
        tc_uuid = UUID(test_case_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid test_case_id format")
        
    try:
        with SessionLocal() as db:
            tc = db.query(TestCase).filter(TestCase.id == tc_uuid).first()
            if not tc:
                raise HTTPException(status_code=404, detail="Test case not found")
                
            tc_data = {
                "title": tc.title,
                "preconditions": tc.preconditions,
                "test_steps": tc.test_steps,
                "expected_result": tc.expected_result
            }
            
        from app.services.agent.bug_report_service import generate_bug_report
        report_content = generate_bug_report(tc_data, payload.actual_result)
        
        return {"report": report_content}
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Error generating bug report: {exc}")
        raise HTTPException(status_code=500, detail=str(exc)) from exc
