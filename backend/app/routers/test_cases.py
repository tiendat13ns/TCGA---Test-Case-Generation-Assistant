from fastapi import APIRouter, HTTPException, Response
from sqlalchemy.exc import SQLAlchemyError
import io

from app.schemas.test_case_schema import GenerateTestCasesResponse, ListTestCasesResponse
from app.services.test_case_generation_service import (
    TestCaseGenerationError,
    generate_test_cases_from_requirement,
    list_test_cases_by_requirement,
)

router = APIRouter(prefix="/api/v1/requirements", tags=["test-cases"])


@router.post(
    "/{requirement_id}/test-cases/generate",
    response_model=GenerateTestCasesResponse,
)
async def generate_test_cases(requirement_id: str):
    try:
        return await generate_test_cases_from_requirement(requirement_id)
    except TestCaseGenerationError as exc:
        raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc


@router.get(
    "/{requirement_id}/test-cases",
    response_model=ListTestCasesResponse,
)
def get_test_cases(requirement_id: str):
    try:
        return list_test_cases_by_requirement(requirement_id)
    except TestCaseGenerationError as exc:
        raise HTTPException(status_code=exc.status_code, detail=str(exc)) from exc
    except SQLAlchemyError as exc:
        raise HTTPException(status_code=500, detail="Database error while loading test cases") from exc


@router.get(
    "/{requirement_id}/test-cases/export",
    responses={
        200: {
            "content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}},
            "description": "Exported Excel file containing test cases.",
        }
    },
)
def export_test_cases(requirement_id: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        # Load test cases from DB
        response_data = list_test_cases_by_requirement(requirement_id)
        tcs = response_data.test_cases
        
        # We need the requirement to get the Feature name (functional_requirement)
        # For simplicity, we can fetch it here or just leave it empty if we don't have the repo
        # Let's import SessionLocal and Requirement to get the feature name
        from app.database import SessionLocal
        from app.models import Requirement
        from uuid import UUID
        
        feature_name = "Unknown Feature"
        try:
            with SessionLocal() as db:
                req = db.get(Requirement, UUID(requirement_id))
                if req and req.functional_requirement:
                    feature_name = req.functional_requirement
        except Exception:
            pass

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        headers = ["Feature", "Test Case ID", "Test Item", "Precondition", "Test Steps", "Test Data", "Expected Output", "Note"]
        ws.append(headers)
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        for idx, tc in enumerate(tcs):
            # Format test steps as a numbered string
            steps_text = "\n".join([f"{i+1}. {s}" for i, s in enumerate(tc.test_steps)]) if tc.test_steps else ""
            
            ws.append([
                feature_name,
                f"TC-{str(idx+1).zfill(2)}",
                tc.title,
                tc.preconditions or "",
                steps_text,
                tc.test_data or "",
                tc.expected_result,
                ""  # Note
            ])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="test_cases_{requirement_id[:8]}.xlsx"'
            }
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
