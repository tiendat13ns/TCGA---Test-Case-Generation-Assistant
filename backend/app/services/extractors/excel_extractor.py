from app.services.extractors.base import BaseExtractor


class ExcelExtractor(BaseExtractor):
    def extract(self, file_path: str) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required to extract XLSX files") from exc

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Could not read XLSX file: {exc}") from exc

        parts: list[str] = []

        for worksheet in workbook.worksheets:
            parts.append(f"===== Sheet: {worksheet.title} =====")
            rows: list[str] = []

            for row in worksheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                line = " | ".join(values).strip()
                if line:
                    rows.append(line)

            parts.append("\n".join(rows))

        workbook.close()
        return "\n\n".join(part for part in parts if part).strip()
